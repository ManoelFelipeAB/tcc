import openpyxl

# Carregar o arquivo Excel
db_2022 = 'C:/Users/agentes/Documents/Projeto HUC/dataset/datasets merged/dez2019_2022/relatorio_2022.xlsx'
workbook = openpyxl.load_workbook(db_2022)

# Selecionar uma planilha (substitua 'Sheet1' pelo nome da planilha)
sheet = workbook['relatorio_2022']

new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

x=1
for row in sheet.iter_rows(min_row=1, values_only=True):
    if x == 1:
        new_sheet.append(row)
        x = 0


# Percorra as linhas da planilha original e copie as linhas que atendem ao critério
for row in sheet.iter_rows(min_row=1, values_only=True):
    first_column_value = row[0]
    if isinstance(first_column_value, int):
        new_sheet.append(row)

# Salve o novo arquivo Excel
new_excel_file_path = 'db_2022.xlsx'
new_workbook.save(new_excel_file_path)

# Feche os arquivos Excel
workbook.close()
new_workbook.close()

print("Linhas filtradas e salvas no novo arquivo.")

print("Linhas removidas com sucesso e novo arquivo salvo.")
